from __future__ import annotations

import json
import struct
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

import app.services.imports as imports_module
from app.services.crypto import CryptoService
from app.services.imports import (
    ImportFormatError,
    ImportLimits,
    ImportParser,
    ImportTooLarge,
)


def crypto() -> CryptoService:
    import base64

    key = base64.b64encode(b"i" * 32).decode()
    return CryptoService.from_secret_values(key, key)


def rotated_crypto() -> CryptoService:
    import base64

    first = base64.b64encode(b"i" * 32).decode()
    second = base64.b64encode(b"j" * 32).decode()
    ring = json.dumps({"active_version": 2, "keys": {"1": first, "2": second}})
    return CryptoService.from_secret_values(ring, ring)


class FakeBlacklist:
    def __init__(self, blocked: set[str] | None = None) -> None:
        self.blocked = blocked or set()
        self.seen: set[str] = set()

    async def matches(self, candidates: set[str]) -> set[str]:
        self.seen = candidates
        return candidates & self.blocked


def xlsx_archive(entries: list[tuple[str, bytes]]) -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        for name, payload in entries:
            archive.writestr(name, payload)
    return output.getvalue()


def synthetic_central_directory(
    actual_entries: int,
    *,
    declared_entries: int,
    comment_bytes: int = 0,
) -> bytes:
    records: list[bytes] = []
    for _ in range(actual_entries):
        record = bytearray(46 + comment_bytes)
        record[:4] = b"PK\x01\x02"
        struct.pack_into("<H", record, 32, comment_bytes)
        records.append(bytes(record))
    central = b"".join(records)
    end = struct.pack(
        "<4s4H2LH",
        b"PK\x05\x06",
        0,
        0,
        declared_entries,
        declared_entries,
        len(central),
        0,
        0,
    )
    return central + end


def mark_first_zip_entry_encrypted(payload: bytes) -> bytes:
    mutated = bytearray(payload)
    local = mutated.find(b"PK\x03\x04")
    central = mutated.find(b"PK\x01\x02")
    assert local >= 0 and central >= 0
    struct.pack_into("<H", mutated, local + 6, struct.unpack_from("<H", mutated, local + 6)[0] | 1)
    struct.pack_into(
        "<H",
        mutated,
        central + 8,
        struct.unpack_from("<H", mutated, central + 8)[0] | 1,
    )
    return bytes(mutated)


@pytest.mark.asyncio
async def test_csv_import_protects_valid_rows_and_removed_list_has_no_plaintext() -> None:
    service = ImportParser(crypto(), FakeBlacklist())
    result = await service.parse(
        "phones.csv",
        BytesIO("手机号\n13800138000\nnot-a-phone\n13800138000\n13900139000\n".encode()),
        size=60,
    )
    assert [item.phone_mask for item in result.valid] == ["138****8000", "139****9000"]
    assert len(result.valid[0].phone_hmac) == 64
    assert result.invalid == 1
    assert result.duplicate == 1
    assert "13800138000" not in result.removed_csv
    assert "not-a-phone" not in result.removed_csv


@pytest.mark.asyncio
async def test_import_enforces_size_and_row_limits_before_persistence() -> None:
    parser = ImportParser(crypto(), FakeBlacklist(), limits=ImportLimits(10, 2))
    with pytest.raises(ImportTooLarge):
        await parser.parse("a.csv", BytesIO(b"1"), size=11)
    with pytest.raises(ImportTooLarge):
        await parser.parse(
            "a.csv",
            BytesIO(b"13800138000\n13900139000\n13700137000\n"),
            size=36,
        )


def test_import_limits_are_built_from_runtime_policy() -> None:
    from app.services.runtime_policy import RuntimePolicy

    limits = ImportLimits.from_policy(
        RuntimePolicy.from_mapping({"import_max_mb": "2", "import_max_rows": "7"})
    )

    assert limits == ImportLimits(2 * 1024 * 1024, 7, 24)


def test_import_expiry_is_built_from_runtime_policy() -> None:
    from app.services.runtime_policy import RuntimePolicy

    limits = ImportLimits.from_policy(
        RuntimePolicy.from_mapping({"import_expire_hours": "6"})
    )

    assert limits.expire_hours == 6


@pytest.mark.asyncio
async def test_blacklist_is_removed_after_hmac_matching() -> None:
    service_crypto = crypto()
    blocked = service_crypto.protect_phone("13800138000").phone_hmac
    result = await ImportParser(service_crypto, FakeBlacklist({blocked})).parse(
        "a.csv",
        BytesIO(b"13800138000\n13900139000\n"),
        size=24,
    )
    assert result.blacklisted == 1
    assert [item.phone_mask for item in result.valid] == ["139****9000"]


@pytest.mark.asyncio
async def test_import_matches_historical_hmac_but_keeps_only_active_tuple() -> None:
    service_crypto = rotated_crypto()
    candidates = service_crypto.hmac_candidates("13800138000")
    blacklist = FakeBlacklist({candidates[1]})

    result = await ImportParser(service_crypto, blacklist).parse(
        "a.csv",
        BytesIO(b"13800138000\n13900139000\n"),
        size=24,
    )

    assert result.blacklisted == 1
    assert candidates[1] in blacklist.seen and candidates[2] in blacklist.seen
    assert all(item.phone_hmac != candidates[1] for item in result.valid)
    assert all(not hasattr(item, "phone_hmacs") for item in result.valid)


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "payload",
    [
        xlsx_archive([("../evil.xml", b"x")]),
        xlsx_archive([("/absolute.xml", b"x")]),
        mark_first_zip_entry_encrypted(xlsx_archive([("xl/workbook.xml", b"x")])),
    ],
)
async def test_xlsx_preflight_rejects_unsafe_paths_and_encrypted_entries(
    monkeypatch: pytest.MonkeyPatch,
    payload: bytes,
) -> None:
    calls: list[object] = []
    monkeypatch.setattr(
        imports_module,
        "load_workbook",
        lambda *args, **kwargs: calls.append((args, kwargs)),
    )

    with pytest.raises(ImportFormatError):
        await ImportParser(crypto(), FakeBlacklist()).parse(
            "phones.xlsx",
            BytesIO(payload),
            size=len(payload),
        )

    assert calls == []


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("payload", "limits"),
    [
        (
            xlsx_archive([("a", b"1"), ("b", b"2"), ("c", b"3")]),
            ImportLimits(max_archive_entries=2),
        ),
        (
            xlsx_archive([("large", b"x" * 101)]),
            ImportLimits(max_entry_uncompressed_bytes=100),
        ),
        (
            xlsx_archive([("a", b"x" * 60), ("b", b"y" * 60)]),
            ImportLimits(max_total_uncompressed_bytes=100),
        ),
        (
            xlsx_archive([("ratio", b"0" * 10_000)]),
            ImportLimits(max_compression_ratio=2),
        ),
    ],
)
async def test_xlsx_preflight_rejects_archive_resource_amplification(
    monkeypatch: pytest.MonkeyPatch,
    payload: bytes,
    limits: ImportLimits,
) -> None:
    calls: list[object] = []
    monkeypatch.setattr(
        imports_module,
        "load_workbook",
        lambda *args, **kwargs: calls.append((args, kwargs)),
    )

    with pytest.raises(ImportTooLarge):
        await ImportParser(crypto(), FakeBlacklist(), limits=limits).parse(
            "phones.xlsx",
            BytesIO(payload),
            size=len(payload),
        )

    assert calls == []


def test_xlsx_preflight_counts_central_records_before_zipfile(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = synthetic_central_directory(257, declared_entries=1)
    monkeypatch.setattr(
        imports_module,
        "ZipFile",
        lambda *_args, **_kwargs: pytest.fail("ZipFile must not be constructed"),
    )

    with pytest.raises(ImportTooLarge, match="条目过多"):
        ImportParser(crypto(), FakeBlacklist()).preflight(
            "phones.xlsx",
            BytesIO(payload),
            size=len(payload),
        )


def test_xlsx_preflight_bounds_central_metadata_before_zipfile(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = synthetic_central_directory(
        2,
        declared_entries=2,
        comment_bytes=32,
    )
    monkeypatch.setattr(
        imports_module,
        "ZipFile",
        lambda *_args, **_kwargs: pytest.fail("ZipFile must not be constructed"),
    )

    with pytest.raises(ImportTooLarge, match="元数据过大"):
        ImportParser(
            crypto(),
            FakeBlacklist(),
            limits=ImportLimits(max_archive_metadata_bytes=100),
        ).preflight(
            "phones.xlsx",
            BytesIO(payload),
            size=len(payload),
        )


def test_xlsx_preflight_rejects_forged_declared_entry_count(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = synthetic_central_directory(2, declared_entries=1)
    monkeypatch.setattr(
        imports_module,
        "ZipFile",
        lambda *_args, **_kwargs: pytest.fail("ZipFile must not be constructed"),
    )

    with pytest.raises(ImportFormatError, match="条目计数不一致"):
        ImportParser(crypto(), FakeBlacklist()).preflight(
            "phones.xlsx",
            BytesIO(payload),
            size=len(payload),
        )


def test_xlsx_preflight_rejects_zip64_before_zipfile(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = struct.pack(
        "<4s4H2LH",
        b"PK\x05\x06",
        0,
        0,
        0xFFFF,
        0xFFFF,
        0xFFFFFFFF,
        0xFFFFFFFF,
        0,
    )
    monkeypatch.setattr(
        imports_module,
        "ZipFile",
        lambda *_args, **_kwargs: pytest.fail("ZipFile must not be constructed"),
    )

    with pytest.raises(ImportFormatError, match="ZIP64"):
        ImportParser(crypto(), FakeBlacklist()).preflight(
            "phones.xlsx",
            BytesIO(payload),
            size=len(payload),
        )


def test_xlsx_preflight_preserves_bounded_zip_comments() -> None:
    source = BytesIO()
    with ZipFile(source, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("xl/workbook.xml", b"workbook")
        archive.comment = b"bounded comment"
    payload = source.getvalue()

    ImportParser(crypto(), FakeBlacklist()).preflight(
        "phones.xlsx",
        BytesIO(payload),
        size=len(payload),
    )


@pytest.mark.asyncio
async def test_valid_small_xlsx_still_imports_after_preflight() -> None:
    source = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["手机号"])
    sheet.append(["13800138000"])
    sheet.append(["13900139000"])
    workbook.save(source)
    source.seek(0)

    result = await ImportParser(crypto(), FakeBlacklist()).parse(
        "phones.xlsx",
        source,
        size=len(source.getvalue()),
    )

    assert [item.phone_mask for item in result.valid] == ["138****8000", "139****9000"]


@pytest.mark.asyncio
async def test_invalid_xlsx_container_is_rejected_before_openpyxl(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    calls: list[object] = []
    monkeypatch.setattr(
        imports_module,
        "load_workbook",
        lambda *args, **kwargs: calls.append((args, kwargs)),
    )

    with pytest.raises(ImportFormatError):
        await ImportParser(crypto(), FakeBlacklist()).parse(
            "phones.xlsx",
            BytesIO(b"not-a-zip"),
            size=9,
        )

    assert calls == []


def test_iter_chunks_bounds_each_database_batch_and_tracks_duplicates_globally() -> None:
    payload = (
        b"phone\n"
        b"13800138000\n"
        b"13900139000\n"
        b"13700137000\n"
        b"13800138000\n"
    )
    chunks = list(
        ImportParser(crypto(), FakeBlacklist()).iter_chunks(
            "phones.csv",
            BytesIO(payload),
            size=len(payload),
            chunk_size=2,
        )
    )

    assert all(len(chunk.valid) + len(chunk.removed) <= 2 for chunk in chunks)
    assert [item.phone_mask for chunk in chunks for item in chunk.valid] == [
        "138****8000",
        "139****9000",
        "137****7000",
    ]
    assert [item.reason for chunk in chunks for item in chunk.removed] == ["duplicate"]
