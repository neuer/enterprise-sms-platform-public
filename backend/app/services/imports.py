"""CSV/XLSX 手机号流式导入、三列保护与无 PII 剔除清单。"""

from __future__ import annotations

import csv
import struct
from collections.abc import Iterator
from dataclasses import dataclass
from io import StringIO, TextIOWrapper
from pathlib import Path, PurePosixPath
from typing import BinaryIO, Protocol
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from app.services.crypto import CryptoService
from app.services.runtime_policy import RuntimePolicy

_EOCD_SIGNATURE = b"PK\x05\x06"
_ZIP64_LOCATOR_SIGNATURE = b"PK\x06\x07"
_CENTRAL_DIRECTORY_SIGNATURE = b"PK\x01\x02"
_EOCD = struct.Struct("<4s4H2LH")
_EOCD_MAX_COMMENT_BYTES = 65_535
_CENTRAL_DIRECTORY_HEADER_BYTES = 46


class ImportTooLarge(ValueError):
    pass


class ImportFormatError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class ImportLimits:
    max_bytes: int = 10 * 1024 * 1024
    max_rows: int = 50_000
    expire_hours: int = 24
    max_archive_entries: int = 256
    max_entry_uncompressed_bytes: int = 32 * 1024 * 1024
    max_total_uncompressed_bytes: int = 64 * 1024 * 1024
    max_compression_ratio: int = 200
    max_archive_metadata_bytes: int = 1024 * 1024

    @classmethod
    def from_policy(cls, policy: RuntimePolicy) -> ImportLimits:
        return cls(
            policy.import_max_mb * 1024 * 1024,
            policy.import_max_rows,
            policy.import_expire_hours,
        )


def _find_eocd(source: BinaryIO, archive_size: int) -> tuple[int, tuple[int, ...]]:
    """在 ZIP 尾部有界查找 EOCD，并拒绝截断或伪造的尾记录。"""

    tail_size = min(archive_size, _EOCD.size + _EOCD_MAX_COMMENT_BYTES)
    source.seek(archive_size - tail_size)
    tail = source.read(tail_size)
    if len(tail) != tail_size:
        raise ImportFormatError("XLSX 压缩包尾记录不完整")
    search_end = len(tail)
    while search_end > 0:
        relative_offset = tail.rfind(_EOCD_SIGNATURE, 0, search_end)
        if relative_offset < 0:
            break
        if relative_offset + _EOCD.size <= len(tail):
            fields = _EOCD.unpack_from(tail, relative_offset)
            comment_size = int(fields[-1])
            if relative_offset + _EOCD.size + comment_size == len(tail):
                absolute_offset = archive_size - tail_size + relative_offset
                return absolute_offset, tuple(int(value) for value in fields[1:])
        search_end = relative_offset
    raise ImportFormatError("XLSX 压缩包缺少有效尾记录")


def _preflight_central_directory(source: BinaryIO, limits: ImportLimits) -> None:
    """在构造 ZipInfo 前有界校验中央目录大小、条目数与结构。"""

    source.seek(0, 2)
    archive_size = source.tell()
    if archive_size > limits.max_bytes:
        raise ImportTooLarge("导入文件超过大小限制")
    if archive_size < _EOCD.size:
        raise ImportFormatError("XLSX 压缩包格式无效")

    eocd_offset, fields = _find_eocd(source, archive_size)
    (
        disk_number,
        central_disk_number,
        entries_on_disk,
        declared_entries,
        central_size,
        central_offset,
        _comment_size,
    ) = fields
    if eocd_offset >= 20:
        source.seek(eocd_offset - 20)
        if source.read(4) == _ZIP64_LOCATOR_SIGNATURE:
            raise ImportFormatError("XLSX 不支持 ZIP64 元数据")
    if (
        entries_on_disk == 0xFFFF
        or declared_entries == 0xFFFF
        or central_size == 0xFFFFFFFF
        or central_offset == 0xFFFFFFFF
    ):
        raise ImportFormatError("XLSX 不支持 ZIP64 元数据")
    if disk_number != 0 or central_disk_number != 0 or entries_on_disk != declared_entries:
        raise ImportFormatError("XLSX 不支持多盘压缩包")
    if declared_entries > limits.max_archive_entries:
        raise ImportTooLarge("XLSX 压缩包条目过多")
    if central_size > limits.max_archive_metadata_bytes:
        raise ImportTooLarge("XLSX 中央目录元数据过大")
    if central_offset + central_size != eocd_offset:
        raise ImportFormatError("XLSX 中央目录边界无效")

    source.seek(central_offset)
    central = source.read(central_size)
    if len(central) != central_size:
        raise ImportFormatError("XLSX 中央目录不完整")
    cursor = 0
    actual_entries = 0
    while cursor < len(central):
        remaining = len(central) - cursor
        if (
            remaining < _CENTRAL_DIRECTORY_HEADER_BYTES
            or central[cursor : cursor + 4] != _CENTRAL_DIRECTORY_SIGNATURE
        ):
            raise ImportFormatError("XLSX 中央目录记录无效")
        filename_size, extra_size, comment_size = struct.unpack_from(
            "<HHH",
            central,
            cursor + 28,
        )
        record_size = (
            _CENTRAL_DIRECTORY_HEADER_BYTES
            + filename_size
            + extra_size
            + comment_size
        )
        if record_size > remaining:
            raise ImportFormatError("XLSX 中央目录记录被截断")
        actual_entries += 1
        if actual_entries > limits.max_archive_entries:
            raise ImportTooLarge("XLSX 压缩包条目过多")
        cursor += record_size
    if actual_entries != declared_entries:
        raise ImportFormatError("XLSX 中央目录条目计数不一致")


@dataclass(frozen=True, slots=True)
class ImportPhone:
    phone_enc: bytes
    phone_hmac: str
    phone_mask: str
    key_version: int
    source_row: int


@dataclass(frozen=True, slots=True)
class RemovedPhone:
    phone_mask: str
    source_row: int
    reason: str


@dataclass(frozen=True, slots=True)
class ImportResult:
    valid: list[ImportPhone]
    removed: list[RemovedPhone]

    @property
    def invalid(self) -> int:
        return sum(item.reason == "invalid" for item in self.removed)

    @property
    def duplicate(self) -> int:
        return sum(item.reason == "duplicate" for item in self.removed)

    @property
    def blacklisted(self) -> int:
        return sum(item.reason == "blacklist" for item in self.removed)

    @property
    def removed_csv(self) -> str:
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["phone_mask", "source_row", "reason"])
        writer.writerows(
            (item.phone_mask, item.source_row, item.reason) for item in self.removed
        )
        return output.getvalue()


@dataclass(frozen=True, slots=True)
class ImportParseChunk:
    """至多一个数据库写批次的受保护号码、HMAC aliases 与掩码剔除项。"""

    valid: tuple[ImportPhone, ...]
    candidates_by_active: dict[str, frozenset[str]]
    removed: tuple[RemovedPhone, ...]


class BlacklistMatcher(Protocol):
    async def matches(self, candidates: set[str]) -> set[str]: ...


class ImportParser:
    def __init__(
        self,
        crypto: CryptoService,
        blacklist: BlacklistMatcher,
        *,
        limits: ImportLimits | None = None,
    ) -> None:
        self.crypto = crypto
        self.blacklist = blacklist
        self.limits = limits or ImportLimits()

    def _rows(self, filename: str, source: BinaryIO) -> Iterator[tuple[int, str]]:
        suffix = Path(filename).suffix.casefold()
        if suffix == ".csv":
            wrapper = TextIOWrapper(source, encoding="utf-8-sig", newline="")
            try:
                for row_no, row in enumerate(csv.reader(wrapper), start=1):
                    yield row_no, row[0].strip() if row else ""
            finally:
                wrapper.detach()
            return
        if suffix == ".xlsx":
            workbook = load_workbook(source, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                if sheet is None:
                    raise ImportFormatError("XLSX 不包含工作表")
                for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    yield row_no, str(row[0]).strip() if row and row[0] is not None else ""
            finally:
                workbook.close()
            return
        raise ImportFormatError("仅支持 csv/xlsx 文件")

    def preflight(self, filename: str, source: BinaryIO, *, size: int) -> None:
        """重型解析前只检查大小、扩展名和 XLSX ZIP 放大风险。"""

        if size < 0 or size > self.limits.max_bytes:
            raise ImportTooLarge("导入文件超过大小限制")
        suffix = Path(filename).suffix.casefold()
        if suffix not in {".csv", ".xlsx"}:
            raise ImportFormatError("仅支持 csv/xlsx 文件")
        if suffix == ".xlsx":
            self._preflight_xlsx(source)
        source.seek(0)

    def _preflight_xlsx(self, source: BinaryIO) -> None:
        """只读检查 XLSX ZIP 元数据，在 openpyxl 解压前限制资源放大。"""

        try:
            source.seek(0)
            _preflight_central_directory(source, self.limits)
            source.seek(0)
            with ZipFile(source) as archive:
                entries = archive.infolist()
                if len(entries) > self.limits.max_archive_entries:
                    raise ImportTooLarge("XLSX 压缩包条目过多")
                total_uncompressed = 0
                total_compressed = 0
                for entry in entries:
                    normalized = entry.filename.replace("\\", "/")
                    path = PurePosixPath(normalized)
                    if (
                        not normalized
                        or normalized.startswith("/")
                        or ".." in path.parts
                        or ":" in path.parts[0]
                        or "\x00" in normalized
                    ):
                        raise ImportFormatError("XLSX 压缩包包含不安全路径")
                    if entry.flag_bits & 0x1:
                        raise ImportFormatError("XLSX 不允许加密条目")
                    if entry.file_size > self.limits.max_entry_uncompressed_bytes:
                        raise ImportTooLarge("XLSX 单个条目解压后过大")
                    entry_ratio_limit = (
                        max(1, entry.compress_size)
                        * self.limits.max_compression_ratio
                    )
                    if entry.file_size > entry_ratio_limit:
                        raise ImportTooLarge("XLSX 单个条目压缩比过高")
                    total_uncompressed += entry.file_size
                    total_compressed += entry.compress_size
                    if total_uncompressed > self.limits.max_total_uncompressed_bytes:
                        raise ImportTooLarge("XLSX 解压后总大小过大")
                if (
                    total_uncompressed
                    > max(1, total_compressed) * self.limits.max_compression_ratio
                ):
                    raise ImportTooLarge("XLSX 总压缩比过高")
        except (ImportFormatError, ImportTooLarge):
            raise
        except (BadZipFile, OSError, RuntimeError, struct.error) as error:
            raise ImportFormatError("XLSX 压缩包格式无效") from error
        finally:
            source.seek(0)

    async def parse(
        self,
        filename: str,
        source: BinaryIO,
        *,
        size: int,
    ) -> ImportResult:
        valid: list[ImportPhone] = []
        removed: list[RemovedPhone] = []
        for chunk in self.iter_chunks(filename, source, size=size):
            blocked_candidates = await self.blacklist.matches(
                set().union(*chunk.candidates_by_active.values())
                if chunk.candidates_by_active
                else set()
            )
            blocked_active = {
                active
                for active, candidates in chunk.candidates_by_active.items()
                if not candidates.isdisjoint(blocked_candidates)
            }
            for item in chunk.valid:
                if item.phone_hmac in blocked_active:
                    removed.append(
                        RemovedPhone(item.phone_mask, item.source_row, "blacklist")
                    )
                else:
                    valid.append(item)
            removed.extend(chunk.removed)
        return ImportResult(valid, removed)

    def iter_chunks(
        self,
        filename: str,
        source: BinaryIO,
        *,
        size: int,
        chunk_size: int = 500,
    ) -> Iterator[ImportParseChunk]:
        """同步分块解析；调用方可逐块查黑名单并批量写库。"""

        if chunk_size < 1 or chunk_size > 2_000:
            raise ValueError("import chunk_size must be 1..2000")
        self.preflight(filename, source, size=size)
        seen: set[str] = set()
        valid: list[ImportPhone] = []
        removed: list[RemovedPhone] = []
        candidates_by_active: dict[str, frozenset[str]] = {}
        data_rows = 0

        def emit() -> ImportParseChunk:
            result = ImportParseChunk(
                tuple(valid),
                dict(candidates_by_active),
                tuple(removed),
            )
            valid.clear()
            candidates_by_active.clear()
            removed.clear()
            return result

        for source_row, value in self._rows(filename, source):
            if source_row == 1 and value.casefold() in {"phone", "mobile", "手机号"}:
                continue
            data_rows += 1
            if data_rows > self.limits.max_rows:
                raise ImportTooLarge("导入文件超过行数限制")
            try:
                protected = self.crypto.protect_phone(value, table="import_phone")
            except ValueError:
                removed.append(RemovedPhone("***********", source_row, "invalid"))
                continue
            if protected.phone_hmac in seen:
                removed.append(
                    RemovedPhone(protected.phone_mask, source_row, "duplicate")
                )
                continue
            seen.add(protected.phone_hmac)
            candidates_by_active[protected.phone_hmac] = frozenset(
                self.crypto.hmac_candidates(value).values()
            )
            valid.append(
                ImportPhone(
                    protected.phone_enc,
                    protected.phone_hmac,
                    protected.phone_mask,
                    protected.key_version,
                    source_row,
                )
            )
            if len(valid) + len(removed) >= chunk_size:
                yield emit()
        if valid or removed:
            yield emit()
